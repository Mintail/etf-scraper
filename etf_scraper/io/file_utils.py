"""
file_utils.py

Internal utilities for reading, writing, and formatting files within the project.

Supported input formats:
    - TXT: Two sections separated by an empty line (columns, then ISINs)
    - CSV: Uses configurable global separator (FileReader.csv_sep)
    - XLSX: Standard Excel file

Supported output formats:
    - CSV
    - XLSX (with optional Excel formatting and adjustable line height)

Classes:
    FileReader:
        Reads input files (TXT, CSV, XLSX) and returns a dictionary with ISINs and column names.
        Allows global control of CSV separator via `FileReader.set_csv_sep()`.

    FileWriter:
        Writes pandas DataFrames to CSV or Excel files.
        Can preserve Excel formatting and adjust row height globally via `FileWriter.set_line_height()`.

Note:
    All classes and functions are intended for internal use only.
"""

import logging
import os
import pandas as pd
import shutil
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet

from etf_scraper.config import get_config
from user_config import WEBSITE, LINE_HEIGHT, MAX_LINE_WIDTH, CSV_SEP

logger = logging.getLogger( __name__ )

class FileReader:
    """
    Class to handle reading input files in TXT, CSV, or XLSX formats.

    Attributes:
        csv_sep (str): Global separator used when reading CSV files.
                       Default value imported from user_config.CSV_SEP.

    Methods:
        set_csv_sep( sep: str ) -> None:
            Changes the global CSV separator for all reads.

        read( input_path: str ) -> dict:
            Detects file type and returns its content as a dictionary containing ISINs and column names.
    """
    csv_sep: str = CSV_SEP

    @classmethod
    def set_csv_sep( cls, sep: str ) -> None:
        """
        Change global CSV separator.
        """
        cls.csv_sep = sep
        logger.info( f"FileReader.csv_sep changed to '{sep}'" )

    @staticmethod
    def read( input_path: str ) -> dict:
        """
        Reads a file and returns its content as a dictionary with isin and columns.
        If len( columns ) < 2, use default columns. 
        
        Automatically detects file type (TXT, CSV, or XLSX).

        Args:
            input_path (str): Path to the input file.

        Returns:
            dict: { "isins": list, "columns": list}

        Raises:
            ValueError: If the file extension is unsupported.
        """
        extension = os.path.splitext( input_path )[ 1 ].lower()
        if extension == ".txt":
            input_processed = FileReader._read_txt( input_path )
        elif extension == ".csv":
            input_processed = FileReader._read_csv( input_path )
        elif extension == ".xlsx":
            input_processed = FileReader._read_excel( input_path )
        else:
            raise ValueError( f"Unsupported input file format: {input_path}" )

        if len( input_processed[ "columns" ] ) < 2:
            input_processed[ "columns" ] = get_config( WEBSITE ).columns_to_fill
        else:
            input_processed[ "columns" ] = FileReader._normalize_columns( input_processed[ "columns" ] )
        
        return input_processed
        

    @staticmethod
    def _read_txt( txt_path: str ) -> dict:
        """
        Reads a TXT file and separates columns and isins based on empty line.
        If no empty line, read only isins only and set columns to [ "ISIN" ].
        """
        data = { "isins": [], "columns": [ "ISIN" ] }
        key  = "columns"
        with open( txt_path, "r" ) as f:
            for line in f:
                if line.strip() == "":
                    key = "isins"
                    continue
                data[ key ].append( line.strip() )
        
        if not data[ "isins" ]:
            data[ "isins" ]   = data[ "columns" ][ 1: ]
            data[ "columns" ] = [ "ISIN" ]
        
        return data

    @staticmethod
    def _read_csv( csv_path: str ) -> dict:
        """
        Reads a CSV file into a dictionary with isins and column names.
        """
        df = pd.read_csv( csv_path, sep = FileReader.csv_sep )
        return { "isins": list(df.iloc[:, 0]), "columns": list(df.columns) }

    @staticmethod
    def _read_excel( excel_path: str ) -> dict:
        """
        Reads an Excel file into a dictionary with isins and column names.
        """
        df = pd.read_excel( excel_path )
        return { "isins": list( df.iloc[ :, 0 ] ), "columns": list( df.columns ) }

    @staticmethod
    def _normalize_columns( columns: list ) -> list:
        """
        Normalize column names matching them against the list of columns to fill.
        """
        columns_to_fill = get_config( WEBSITE ).columns_to_fill
        columns_norm = []
        for column in columns:
            original = column
            column_clean = column.strip().lower()
            # Try to match with known fillable columns
            matched = next(
                ( ref for ref in columns_to_fill if ref.lower() == column_clean ),
                None
            )
            normalized = matched if matched else original
            # Avoid duplicates while preserving order
            if normalized not in columns_norm:
                columns_norm.append( normalized )

            if not matched :
                logger.warning( "The following column will not be filled: '%s'", column )

        return columns_norm


class FileWriter:
    """
    Class to handle writing pandas DataFrames to CSV or Excel files.

    This class supports both plain exports (CSV or Excel) and Excel exports 
    that reuse an existing file as a formatting template. Formatting adjustments 
    such as column width, row height, and text alignment are handled automatically.

    Attributes:
        line_height (int): Global line height multiplier for Excel rows.
                           Default value imported from user_config.LINE_HEIGHT.

    Methods:
        set_line_height(height: int) -> None:
            Changes the global line height multiplier for Excel outputs.

        write(df, output_path, use_input_format = False, input_path = None, override_input = False) -> None:
            Writes a DataFrame to a file, optionally preserving Excel formatting from an existing template.
            Invalid template references are ignored gracefully and logged as warnings.
    """
    line_height: int = LINE_HEIGHT

    @classmethod
    def set_line_height( cls, height: int) -> None:
        """
        Change global Excel line height multiplier.
        """
        if height <= 0:
            raise ValueError( "line_height must be positive" )
        cls.line_height = height
        logger.info( f"FileWriter.line_height changed to {height}" )

    @staticmethod
    def write(
        df: pd.DataFrame,
        output_path: str,
        use_input_format: bool = False,
        input_path: str = None,
        override_input: bool = False
    ) -> None:
        """
        Writes a pandas DataFrame to a CSV or Excel file.

        Behavior:
            - If 'use_input_format' is False:
                The DataFrame is written normally.
            - If 'use_input_format' is True:
                Attempts to reuse formatting from 'input_path'. If the file does not exist,
                is not an Excel file, or if extensions are incompatible, the operation
                falls back to a standard write. Each fallback is logged as a warning.

        Args:
            df (pd.DataFrame): The DataFrame to write.
            output_path (str): Destination file path (.csv or .xlsx).
            use_input_format (bool, optional): 
                If True, reuse formatting from 'input_path' (Excel only).
                Defaults to False.
            input_path (str, optional): 
                Path to an existing Excel file used for formatting when 'use_input_format' is True.
            override_input (bool, optional): 
                If True and 'input_path' == 'output_path', the file is overwritten directly.
                If False, a copy is created to avoid data loss.
                Defaults to False.

        Raises:
            ValueError: If the output format is unsupported.
        
        Notes:
            Invalid or missing input_path are logged as warnings and the export continues using the default formatting.
        """

        extension = os.path.splitext( output_path )[ 1 ].lower()
        if use_input_format:
            if not input_path:
                logger.warning( f"use_input_path is True but input_path is not provided: ignoring input_format." )
                use_input_format = False
            elif not os.path.exists( input_path ):
                logger.warning( f"Input file at '{input_path}' does not exist: ignoring input_format." )
                use_input_format = False
            elif os.path.splitext( input_path )[ 1 ].lower() != ".xlsx":
                logger.warning( f"Input file at'{input_path}' is not .xlsx: ignoring input_format." )
                use_input_format = False
            elif extension != ".xlsx":
                logger.warning( f"Output file at'{input_path}' is not .xlsx: ignoring input_format." )
                use_input_format = False

        if extension == ".csv":
            df.to_csv( output_path, index = False )
        elif extension == ".xlsx":
            FileWriter._write_excel( df, output_path, use_input_format, input_path, override_input )
        else:
            raise ValueError( f"Unsupported output file format: {output_path}" )

        logger.info( f"DataFrame successfully written to '{output_path}'" )

    @staticmethod
    def _write_excel(
        df: pd.DataFrame,
        output_path: str,
        use_input_format: bool = False,
        input_path: str = None,
        override_input: bool = False
    ) -> None:
        """
        Writes DataFrame to an Excel file with optional formatting.
        """
        if use_input_format:
            if input_path != output_path:
                shutil.copy( input_path, output_path )
            elif not override_input:
                input_path_name      = os.path.splitext( output_path )[ 0 ].lower()
                input_path_extension = os.path.splitext( output_path )[ 1 ].lower()
                input_path_copy      = input_path_name + "_copy." + input_path_extension
                shutil.copy( input_path, input_path_copy )
            wb = load_workbook( output_path )
            ws = wb.active
            # Update only values (avoid rewriting all styles)
            for r_idx, row in enumerate( df.itertuples( index = False ), start = 2 ):
                for c_idx, value in enumerate( row, start = 1 ):
                    ws.cell( row = r_idx, column = c_idx ).value = value
        else:
            # For large files, use pandas ExcelWriter with openpyxl engine
            with pd.ExcelWriter( output_path, engine = 'openpyxl' ) as writer:
                df.to_excel( writer, index = False )
            wb = load_workbook( output_path )
            ws = wb.active

        # Adjust formatting (column widths + row heights + center cells)
        FileWriter._set_columns_width( ws )
        FileWriter._set_rows_height( ws )
        FileWriter._center_cells( ws )
        wb.save( output_path )
    
    @staticmethod
    def _set_columns_width( ws: Worksheet ) -> None:
        """
        Adjust column widths to fit content, minimum width is 10.
        """ 
        for column_cells in ws.columns:
            max_length = 0
            max_width  = MAX_LINE_WIDTH - 2
            col_letter = get_column_letter( column_cells[ 0 ].column )
            for cell in column_cells:
                if cell.value:
                    length_cell = max( [ len( line ) for line in str( cell.value ).split( "\n" ) ] )
                    if length_cell > max_width:
                        cell.value  = FileWriter._split_cell( str( cell.value ), max_width )
                        length_cell = max( [ len( line ) for line in str( cell.value ).split( "\n" ) ] )
                    max_length  = max( max_length, length_cell )
            ws.column_dimensions[ col_letter ].width = max( 10, max_length + 2 )

    @staticmethod
    def _split_cell( value: str, max_width: int ) -> str:
        """
        Split value in blank space so that the length does not exceed max_width.
        """
        words = value.split( " " )
        lines = []
        current_line = ""
        for word in words:  
            while len( word ) > max_width: # Cut word if its length > max_width
                if current_line:
                    lines.append( current_line )
                    current_line = ""
                lines.append( word[ :max_width ] )
                word = word[ max_width: ]

            if not current_line:
                current_line = word
            elif len( current_line ) + len( word ) < max_width:
                current_line += " " + word
            else:
                lines.append( current_line )
                current_line = word
        
        if current_line:
            lines.append( current_line )
        
        return "\n".join( lines )

    @staticmethod
    def _set_rows_height( ws: Worksheet ) -> None:
        """
        Adjusts Excel row heights based on cell content and global line_height.
        """
        for row_cells in ws.iter_rows():
            max_height = 1
            for cell in row_cells:
                if cell.value:
                    lines = str( cell.value ).count( "\n" ) + 1
                    max_height = max( max_height, lines * FileWriter.line_height )
            ws.row_dimensions[ cell.row ].height = max_height
    
    @staticmethod
    def _center_cells( ws: Worksheet ) -> None:
        """
        Centers the text in each cell.
        """   
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment( horizontal = "center", vertical = "center" )